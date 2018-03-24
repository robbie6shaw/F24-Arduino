#Program to extract form text file serial monitor data to an excel sheet. Version Built for Battery Tester Program
#10 March 2018
#Alpha Unstable

import sys
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

cycleCount = 3 
variableCount = 7

rawLocation = 'F:\\F24\\Arduino Code\\Python Extraction Program\\RawDataExample.txt'
processedLocation = 'F:\\F24\\Arduino Code\\Python Extraction Program\\Battery Data.xlsx'

dataRead = open(rawLocation)
rawData = dataRead.read()
rawDataList = rawData.split('#')

print('Please Enter name of Test:')
sheetName = input()

wb = openpyxl.load_workbook(processedLocation)
wa = wb.create_sheet(title=sheetName)
wa['A1']='Cycle Count'
wa['B1']='Unloaded Voltage'
wa['C1']='Unloaded Current'
wa['D1']='Inital Loaded Voltage'
wa['E1']='Inital Loaded Current'
wa['F1']='Final Loaded Voltage'
wa['G1']='Final Loaded Current'

for i in range(variableCount):
    column = get_column_letter(i+1)
    wa.column_dimensions[column].width=20

for i in range(cycleCount):
    oneTimeStr =  rawDataList[i]
    oneTimeList = oneTimeStr.split('\n')

    for j in range(variableCount):

        value = oneTimeList [j]
        valueList = value.split('=')

        if 't' in valueList:
            timeStamp = valueList[1]
            cell = 'A'+str(i+2)
            wa[cell] = float(timeStamp) 
        
        elif 'a' in valueList:
            unloadedVoltage = valueList[1]
            cell = 'B'+str(i+2)
            wa[cell] = float(unloadedVoltage)
            
        elif 'b' in valueList:
            unloadedCurrent = valueList[1]
            cell = 'C'+str(i+2)
            wa[cell] = float(unloadedCurrent)

        elif 'c' in valueList:
            initalLoadedVoltage = valueList[1]
            cell = 'D'+str(i+2)
            wa[cell] = float(initalLoadedVoltage)

        elif 'd' in valueList:
            intialLoadedCurrent =  valueList[1]
            cell = 'E'+str(i+2)
            wa[cell] = float(intialLoadedCurrent)

        elif 'e' in valueList:
            finalLoadedVoltage = valueList[1]
            cell = 'F'+str(i+2)
            wa[cell] = float(finalLoadedVoltage)

        elif 'f' in valueList:
            finalLoadedCurrent = valueList[1]
            cell = 'G'+str(i+2)
            wa[cell] = float(finalLoadedCurrent)
    

wb.save(processedLocation)
print('Finished')
